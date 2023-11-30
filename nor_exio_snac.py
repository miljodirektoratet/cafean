# -*- coding: utf-8 -*-
# %% [markdown]
#  # Hybrid SNAC model adopted for Norway SSB data
#
# Top level script for calculation Norwegian carbon footprints
# based on the hybrid-SNAC approach, coupling EXIOBASE 3
# with official data from SSB Norway
#

# %% [markdown]
# Copyright (C) 2023  XIO Sustainability Analytics, Inc
# 
# Written by
# 
# - Richard Wood
# - Konstantin Stadler
# 
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
# 
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
# 
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

# %% [markdown]
# ## Usage Notes
#
# This file requires the extract of import/final demand multipliers from EXIOBASE.
# These are included in the repository (./data/exio_no_bp_raw.xlsx), but can 
# also reproduced using the ./exio_prepper.py script.

# %% [markdown]
# Import of required external packages. These are all available through pip/conda/mamba install.

# %%
import numpy as np
import openpyxl as opx

import pandas as pd
import pymrio

# %% [markdown]
# Python internal packages (this don't need to be installed, part of standard python)

# %%
from pathlib import Path

# %% [markdown]
# ## Settings

# %% [markdown]
# ### Parameters
# Run the model for the following years:

# %%
# range command is exclusive of the last number, so we add 1
years = list(range(2012, 2021))
# for only running one year we can do
# years: list[int] = [2020]

# %% [markdown]
# ### Path definitions

# %%
# Set the work path to the directory where this script is located
# and if this is not available to the current working directory
try:
    work_path = Path(__file__).parent.absolute()  # when running as script
except NameError:
    work_path = Path.cwd()

data_path: Path = work_path / "data"

output_path: Path = work_path / "results"
output_path.mkdir(parents=True, exist_ok=True)

# %% [markdown]
# ### Specifying/reading data sources
# This section defines all data sources needed for the calculation.
# Smaller dataset are read already here, larger datasets are read
# at the places they are needed.

# %% [markdown]
# The sector matching file provides the bridge from the
# EXIOBASE sector classification to the Norwegian official
# classification.

# %%
sector_matching = pd.read_excel(
    data_path / "sector_matching.xlsx",
    sheet_name="EXIOBASEtoNO65",
    index_col=[1, 2, 3],
    header=[1, 2, 3]).iloc[:, 1:] # get rid of the summation column

sector_matching.index.names = ["nr", "code", "name"]
sector_matching.columns.names = ["nr", "name", "code"]

# %% [markdown]
# The sector list file contains the names used for
# parsing the sectors in the emission data. It is also
# used for then renaming the sectors to the names
# used in the official economic data.
# *NOTE:* The official economic data is provided in
# industry by industry (ixi) format, but the names
# are "product names". We stay consistent to that
# and also use product names throughout the script.

# %%
sector_list = pd.read_excel(data_path / "sector_list.xlsx")

# %% [markdown]
# The emission characterization file contains
#   1. The conversion factors from EXIOBASE emission units/names to Norwegian emission units/names
#   2. The characterization of GHG to GWP100
# *NOTE: *: All unit conversion for emission data happens in that file!
# These units must match the units provided by the EXIOBASE and Norwegian emission data.
# Double-check these when updating the data!


# %%
exio_stressor_to_nor = pd.read_excel(
    data_path / "emis_conv_char.xlsx", sheet_name="exio_nor_conv", index_col=0)
charact_ghg = pd.read_excel(
    data_path / "emis_conv_char.xlsx", sheet_name="nor_char", index_col=0)

# %% [markdown]
# The official Norwegian emission data file.

# %%
nor_emission_file = data_path / "AEA Questionnaire_2023_Norway2023_til footprint.xlsm"

# %% [markdown]
# Get the exchange rates per year.
# These are currently manually inserted from Eurostat official data.
# Other sources can be used, just updating the numbers in the file.

# %%
exchange_euro_nok = pd.read_csv(data_path / "exchange_rates.csv", sep="\t", index_col=0)

# %% [markdown]
# The Norwegian IO data files

# %%
nor_io_dom_files = {
    int(file.stem[-4:]): file for file in data_path.glob("iot_1850_*.xlsx")}
nor_io_imp_files = {
    int(file.stem[-4:]): file for file in data_path.glob("iot_1950_*.xlsx")}

# %% [markdown]
# Then we specify the file with the extracted EXIOBASE data.
# This needs to be prepared for each new EXIOBASE version.
# The utility script ./exio_prepper.py can be used for this.
# For version 3.8.2, the file is already provided in the data folder.

# %%
exio_extract_file: Path = data_path / "exio_no_bp_raw.xlsx"

# %% [markdown]
# Last we define the emissions of interest.
# These needs to be the name of the sheets in the emission file.

# %%
emis_in_nor_data = ["CO2", "Biomass CO2", "N2O", "SF6_NF3", "CH4", "HFC", "PFC"]


# %% [markdown]
# ## Helper functions

# %% [markdown]
# Here we define a function used throughout the script to check the data
# and bring it to a consistent format.


# %%
def df_checks(df, name, report_nan=False, report_inf=True, report_str=True):
    """Checks all non-float columns, report and fix

    Checks and fixes:

        - dtypes (cast to float if possible)
        - NaN (set to 0.0)
        - Inf (set to 0.0)
        - strings (set to 0.0)

    Parameters
    ----------
    df : pd.DataFrame or pd.Series
        The dataframe to check
    name : str
        The name of the dataframe, used for reporting.
    report_nan : bool, optional
        Whether to report NaN values, by default False
    report_inf : bool, optional
        Whether to report Inf values, by default True
    report_str : bool, optional
        Whether to report string values, by default True

    Returns
    -------
    df : pd.DataFrame or pd.Series, depending on input
    """
    if type(df) == pd.Series:
        df = df.to_frame()
        return_back_to_series = True
    else:
        return_back_to_series = False

    for col in df.columns:
        if df[col].dtype in (float, np.float64):
            pass
        elif df[col].dtype == int:
            # int can be just cast to float
            df[col] = df[col].astype(float)
        elif df[col].dtype == object:
            # check for strings in column and report these
            row_str = df[col].apply(lambda x: isinstance(x, str))
            if row_str.any():
                for row in df[row_str].index:
                    if report_str:
                        print(
                            f"WARNING: Table {name} - Column {col} contains string {df.loc[row, col]}. Set to 0.0."
                        )
                    df.loc[row, col] = 0.0
            df[col] = df[col].astype(float)
        else:
            print(f"WARNING: {name} - {col} has unhandled dtype {df[col].dtype}.")

        # check for NaN and Inf
        if df[col].isna().any():
            if report_nan:
                print(
                    f"WARNING: Table {name} - Column {col} contains NaN values. Set to 0.0."
                )
            df[col].fillna(0.0, inplace=True)
        if np.isinf(df[col].values).any():
            if report_inf:
                print(
                    f"WARNING: Table {name} - Column {col} contains Inf values. Set to 0.0."
                )
            df[col][np.isinf(df[col].values)] = 0.0

    if return_back_to_series:
        df = df.squeeze()

    return df


# %% [markdown]
# ## Read the official Norwegian emission data


# %% [markdown]
# ### Function for reading the official Norwegian emission data

# %%
def get_nor_emissions(emission_type, years, emission_file, industry_codes):
    """Read the official Norwegian emission data for a given emission type.

    Parameters
    ----------
    emission_type : str
        The emission type to read.
    years : list[int]
        The years to read.
    emission_file : Path
        The path to the emission file.
    industry_codes : list[str]
        The industry codes to read.

    Returns
    -------
    emis_nor : pd.DataFrame
        The emission data for the given emission type.
    emis_nor_hhld : pd.DataFrame
        The household emission data for the given emission type.
    """

    raw_emission_data = pd.read_excel(emission_file, sheet_name=emission_type, header=3)

    emis_nor = pd.DataFrame(index=industry_codes, columns=years)

    year_columns = [year for year in years if year in raw_emission_data.columns]

    slice_hhld = slice(86, 90)
    name_index_hhld_total = (86, 1)
    name_index_hhld_breakdown = (slice(87, 90), 3)

    for icode in industry_codes:
        indx = raw_emission_data[
            raw_emission_data["Air emissions by industry"] == icode
        ].index
        if not indx.empty:
            indx = indx[0]
            data = raw_emission_data.loc[indx, year_columns].values
            emis_nor.loc[icode, year_columns] = data

    hhld_total_name = (
        raw_emission_data.iloc[*name_index_hhld_total].split("\n")[1].strip()
    )
    hhld_breakdown_name = (
        raw_emission_data.iloc[*name_index_hhld_breakdown].str.strip("-").str.strip()
    )
    hhld_names = [hhld_total_name] + hhld_breakdown_name.tolist()

    emis_nor_hhld = raw_emission_data.loc[:, year_columns].iloc[slice_hhld, :]
    emis_nor_hhld.index = hhld_names

    emis_nor = df_checks(
        emis_nor,
        "Norwegian industry emissions",
        report_nan=True,
        report_inf=True,
        report_str=True,
    )
    emis_nor_hhld = df_checks(
        emis_nor_hhld,
        "Norwegian household emissions",
        report_nan=True,
        report_inf=True,
        report_str=True,
    )

    # Put data from row c19 into row c21 - to be consistent with IO data.
    emis_nor.loc["C21"] = emis_nor.loc["C19"]
    emis_nor.loc["C19"] = 0

    emis_nor_hhld.index.name = "emission_type"
    emis_nor_hhld.columns.name = "activity"
    emis_nor.index.name = "emission_type"

    return emis_nor, emis_nor_hhld


# %% [markdown]
# Next, we read the emission data into two
# dictionaries (one for industry and one for households) and then
# convert these to yearly emission data with emission type as rows
# and industry as columns.

# %%
emis_type_dict = {}
emis_type_dict_hhld = {}

for emission in emis_in_nor_data:
    print(f"PROCESSING emission data for {emission}.")
    emis_type_dict[emission], emis_type_dict_hhld[emission] = get_nor_emissions(
        emission_type=emission,
        years=years,
        emission_file=nor_emission_file,
        industry_codes=sector_list.IndustryCode.dropna(),
    )

emis_col_ind = {}
emis_col_hhld = {}

for year in years:
    emis_col_ind[year] = pd.DataFrame(
        data=[emis_type_dict[emission][year].T for emission in emis_in_nor_data],
        index=emis_in_nor_data,
    )
    emis_col_hhld[year] = pd.DataFrame(
        data=[
            emis_type_dict_hhld[emission][year].to_dict()
            for emission in emis_in_nor_data
        ],
        index=emis_in_nor_data,
    )

# %% [markdown]
# Then we characterize the Norwegian emission data

# %%
emis_nor_ind = {}
emis_nor_hhld = {}
for year in years:
    ghg_name = charact_ghg.impact.unique()
    if len(ghg_name) > 1:
        raise NotImplementedError(
            "More than one characerized impact in characterization factors."
        )
    else:
        ghg_name = ghg_name[0]

    # First we convert to GHGeq based on characterization factors.
    # These are not stored currently, but we could save them if needed.
    emis_ind_co2eq = emis_col_ind[year].multiply(charact_ghg.factor, axis=0)
    emis_hhld_co2eq = emis_col_hhld[year].multiply(charact_ghg.factor, axis=0)

    # Then we sum got GHG totals
    emis_ind_co2eq.loc[ghg_name, :] = emis_ind_co2eq.sum(axis=0).fillna(0)
    emis_hhld_co2eq.loc[ghg_name, :] = emis_hhld_co2eq.sum(axis=0).fillna(0)

    # And add them to the native (non CO2eq) format data
    emis_nor_ind[year] = emis_col_ind[year].copy()
    emis_nor_hhld[year] = emis_col_hhld[year].copy()
    emis_nor_ind[year].loc[ghg_name, :] = emis_ind_co2eq.loc[ghg_name, :]
    emis_nor_hhld[year].loc[ghg_name, :] = emis_hhld_co2eq.loc[ghg_name, :]
    emis_nor_hhld[year].index.names = ["emission"]
    emis_nor_hhld[year].columns.names = ["hhld_component"]

# %% [markdown]
# Then we rename columns to match the Norwegian economic sector names

# %%
rename_dict = (
    sector_list.loc[:, ["IndustryCode", "Product"]]
    .dropna(how="any")
    .set_index("IndustryCode")
    .squeeze()
    .to_dict()
)

for year in years:
    emis_nor_ind[year] = emis_nor_ind[year].rename(columns=rename_dict)
    # strip white space since "Coke and refined..." sector has white space
    # in the official data
    emis_nor_ind[year].columns = emis_nor_ind[year].columns.str.strip()
    emis_nor_ind[year].index.name = "emission"
    emis_nor_ind[year].columns.name = "sector"

# %% [markdown]
# Next we make a summary dataframe holding the units for each emission type.
# We also order that and use this order during saving the results.

# %%
emis_nor_unit = charact_ghg.loc[:, "stressor_unit"]
emis_nor_unit.name = "unit"
emis_nor_unit.loc[ghg_name] = charact_ghg.loc[:, "impact_unit"].unique()[0]
emis_nor_unit.index.name = "emission"
# order the rows by impact and then the order of stressor in charact_ghg
_sort_emis = charact_ghg.impact.unique().tolist() + charact_ghg.index.unique().tolist()
emis_nor_unit = emis_nor_unit.loc[_sort_emis]

# %% [markdown]
# ## IO reading functions

# %% [markdown]
# Function for reading one year from the Norwegian IO tables.
# The official IO tables have a quite complex structure, so
# the function below is quite complex as well. There are
# several "magic strings and numbers" which refer to specific
# cells in the excel files. These will need to be
# updated if the structure of the excel files changes.


# %%
def get_nor_IO(file_dom, file_imp, emis_nor_df):
    """Read the official Norwegian IO tables and calculate L, x, SL.

    Parameters
    ----------

    file_dom : Path
        The path to the domestic IO table.

    file_imp : Path
        The path to the import IO table.

    emis_nor_df : pd.DataFrame
        The emission data for the specific year.

    Returns
    -------
    product_names : pd.DataFrame

    nor_io : dict
        IO tables in a dictionary.
    """
    # Sheet definitions
    import_sheet = "1950"
    domestic_sheet = "1850"

    # Cell definitions
    cells_product_names = "D29:D93"
    cells_Ydom_name = "BS27:CG27"
    cells_Yimp_name = "BS27:CG27"
    # The final use string is one row above the others ¯\_(ツ)_/¯
    cell_Ydom_final_use_name = "CG26"
    cell_Yimp_final_use_name = "CG26"

    cells_Zdom = "E29:BQ93"
    cells_Ydom = "BS29:CG93"

    cells_Zimp = "E29:BQ93"
    cells_Yimp = "BS29:CG93"

    cells_x_tot = "CH29:CH93"

    def cell_reader(wb_sheet, cell_range):
        """Helper function for reading cell ranges"""
        data = []
        for row in wb_sheet[cell_range]:
            data.append([cell.value for cell in row])
        df = pd.DataFrame(data)
        return df

    wb_dom = opx.load_workbook(file_dom, data_only=True)
    assert (
        domestic_sheet in wb_dom.sheetnames
    ), f"Sheet {domestic_sheet} not found in file_dom, structure changed"

    prod_names = (
        cell_reader(wb_dom[domestic_sheet], cells_product_names).iloc[:, 0].str.strip()
    )

    Zdom = cell_reader(wb_dom[domestic_sheet], cells_Zdom)
    Zdom.columns = prod_names
    Zdom.columns.name = "sector"
    Zdom.index = prod_names
    Zdom.index.name = "sector"
    Zdom = df_checks(Zdom, "Zdom")

    Ydom = cell_reader(wb_dom[domestic_sheet], cells_Ydom)
    Ydom_names = cell_reader(wb_dom[domestic_sheet], cells_Ydom_name).iloc[0, :]
    Ydom_names.iloc[-1] = wb_dom[domestic_sheet][cell_Ydom_final_use_name].value
    # strip white space and carriage return
    Ydom_names = [name.strip() for name in Ydom_names]
    Ydom.columns = Ydom_names
    Ydom.columns.name = "category"
    Ydom.index = prod_names
    Ydom.index.name = "sector"
    Ydom = df_checks(Ydom, "Ydom")

    xdom = cell_reader(wb_dom[domestic_sheet], cells_x_tot).fillna(0).iloc[:, 0]
    xdom.index = prod_names
    xdom.index.name = "sector"
    xdom = df_checks(xdom, "xdom")

    wb_dom.close()

    wb_imp = opx.load_workbook(file_imp, data_only=True)
    assert (
        import_sheet in wb_imp.sheetnames
    ), f"Sheet {import_sheet} not found in file_dom, structure changed"

    prod_names_imp = (
        cell_reader(wb_imp[import_sheet], cells_product_names).iloc[:, 0].str.strip()
    )

    assert prod_names_imp.equals(
        prod_names
    ), "Product names in import and domestic IO tables do not match"

    Zimp = cell_reader(wb_imp[import_sheet], cells_Zimp).fillna(0)
    Zimp.columns = prod_names
    Zimp.columns.name = "sector"
    Zimp.index = prod_names
    Zimp.index.name = "sector"
    Zimp = df_checks(Zimp, "Zimp")

    Yimp_raw = cell_reader(wb_imp[import_sheet], cells_Yimp).fillna(0)
    Yimp_names = cell_reader(wb_imp[import_sheet], cells_Yimp_name).iloc[0, :]
    Yimp_names.iloc[-1] = wb_imp[import_sheet][cell_Yimp_final_use_name].value
    # strip white space and carriage return
    Yimp_names = [name.strip() for name in Yimp_names]
    Yimp_raw.columns = Yimp_names
    Yimp_raw.columns.name = "category"
    Yimp_raw.index = prod_names
    Yimp_raw.index.name = "sector"
    Yimp_raw = df_checks(Yimp_raw, "Yimp")

    ximp = cell_reader(wb_imp[import_sheet], cells_x_tot).fillna(0).iloc[:, 0]
    ximp.index = prod_names
    ximp.index.name = "sector"
    ximp = df_checks(ximp, "ximp")

    wb_imp.close()

    nor_io_tot = {}

    nor_io_tot["Zdom"] = Zdom

    nor_io_tot["Ydom_exp"] = Ydom.loc[:, "Exports fob (2)"]

    # Full domestic final demand matrix without aggregates
    # The list below shows all column headers, the ones not used
    # are commented out - just for manual control.
    nor_io_tot["Ydom_br"] = Ydom.loc[
        :,
        [
            "Final consumption expenditure by households",
            "Final consumption expenditure by non-profit organisations serving households (NPISH)",
            "Final consumption expenditure by government",
            # 'Final consumption expenditure',
            "Gross fixed capital formation",
            # 'Changes in valuables (1)',
            "Changes in inventories",
            #'Changes in inventories and valuables',
            #'Gross capital formation',
            #'Exports intra EU fob (1)',
            #'Exports fob to members of the euro area (1)',
            #'Exports fob to non-members of the euro area (1)',
            #'Exports extra EU fob (1)',
            "Exports fob (2)",
            #'Final uses']   # Excluding final uses as these are recalculated
        ],
    ]

    # Domestic final demand (excluding exports)
    # nor_io_tot["Ydom"] = Ydom.loc[:, "Final uses"] - nor_io_tot["Ydom_exp"]

    # Setting negatives to zero, important for GCF and changes in valueables
    nor_io_tot["Ydom_br"][nor_io_tot["Ydom_br"] < 0] = 0

    nor_io_tot["Ydom"] = nor_io_tot["Ydom_br"].sum(axis=1) - nor_io_tot["Ydom_exp"]

    # Import flow matrix
    nor_io_tot["Zimp"] = Zimp

    # Import-exports (re-exports)
    nor_io_tot["Yimp_exp"] = Yimp_raw.loc[:, "Exports fob (2)"]

    nor_io_tot["Yimp_raw"] = Yimp_raw

    # Full import final demand matrix without aggregates
    # The list below shows all column headers, the ones not used
    # are commented out - just for manual control.
    nor_io_tot["Yimp_br"] = Yimp_raw.loc[
        :,
        [
            "Final consumption expenditure by households",
            "Final consumption expenditure by non-profit organisations serving households (NPISH)",
            "Final consumption expenditure by government",
            # 'Final consumption expenditure',
            "Gross fixed capital formation",
            # 'Changes in valuables 1)',
            "Changes in inventories",
            # 'Changes in inventories and valuables',
            # 'Gross capital formation',
            # 'Exports intra EU fob (1)',
            # 'Exports fob to members of the euro area (1)',
            # 'Exports fob to non-members of the euro area (1)',
            # 'Exports extra EU fob (1)',
            "Exports fob (2)",
            # 'Final uses at basic prices',
        ],
    ]

    # Setting negatives to zero, important for GCF and changes in valueables
    nor_io_tot["Yimp_br"][nor_io_tot["Yimp_br"] < 0] = 0

    nor_io_tot["Yimp"] = nor_io_tot["Yimp_br"].sum(axis=1) - nor_io_tot["Yimp_exp"]

    # Recalculating x considering removed negatives
    nor_io_tot["x"] = (
        nor_io_tot["Zdom"].sum(axis=1) + nor_io_tot["Ydom"] + nor_io_tot["Ydom_exp"]
    )

    nor_io_tot["Adom"] = pymrio.calc_A(nor_io_tot["Zdom"], nor_io_tot["x"])
    nor_io_tot["Aimp"] = pymrio.calc_A(nor_io_tot["Zimp"], nor_io_tot["x"])
    nor_io_tot["Ldom"] = pymrio.calc_L(nor_io_tot["Adom"])

    nor_io_tot["Sdom"] = pymrio.calc_S(emis_nor_df, nor_io_tot["x"])

    nor_io_tot["SLdom"] = pymrio.calc_M(S=nor_io_tot["Sdom"], L=nor_io_tot["Ldom"])
    nor_io_tot["SLdom"]

    return nor_io_tot


# %% [markdown]
# ## Main coupling and calculation step

# %% [markdown]
# First we define several function for coupling and calculations


# %%
def snac_coupling(Aimp, Ldom, Qimp_mrio):
    """Coupling of MRIO based multipliers with domestic IO tables

    Parameters
    ----------

    Aimp: pd.DataFrame
        Import A matrix
    Ldom: pd.DataFrame
        Domestic L matrix
    Qimp_mrio: pd.DataFrame
        Import multipliers from MRIO

    Returns
    -------
    dict with
        QAimp: pd.DataFrame
            QAimp matrix
        QAimpLdom: pd.DataFrame
            QAimpLdom matrix
    """
    QAimp = Qimp_mrio @ Aimp
    QAimpLdom = QAimp @ Ldom

    return dict(QAimp=QAimp, QAimpLdom=QAimpLdom)


# %%
def snac_findem_breakdown(
    SLdom_diag, QAimp_Ldom_diag, Q_imp_exio_diag, Ydom_br, Yimp_br
):
    """Breakdown of the total fp into final demand categories

    Components: domestic_footprint, import_footprint, total

    Parameters
    ----------
    SLdom_diag: pd.DataFrame
        Diagonalized SLdom matrix for one stressor/gas/impact
    QAimp_Ldom_diag: pd.DataFrame
        Diagonalized QAimpLdom matrix for one stressor/gas/impact
    Q_imp_exio_diag: pd.DataFrame
        Diagonalized Q_imp_exio matrix for one stressor/gas/impact
    Ydom_br: pd.DataFrame
        domestic_footprint final demand matrix without aggregates
    Yimp_br: pd.DataFrame
        Import final demand matrix without aggregates

    Returns
    --------
    pd.Series

        index: pd.MultiIndex

            level 0: component
                domestic_footprint, import_footprint, total_footprint
            level 1: sector
                sector names
            level 2: category
                final demand category


    """

    dom_breakdown = SLdom_diag @ Ydom_br
    imp_breakdown = QAimp_Ldom_diag @ Ydom_br + Q_imp_exio_diag @ Yimp_br
    tot_footprint_breakdown = dom_breakdown + imp_breakdown

    index_names = ["component"] + list(dom_breakdown.index.names)

    ds = pd.concat(
        [dom_breakdown, imp_breakdown, tot_footprint_breakdown],
        keys=["domestic_footprint", "import_footprint", "total_footprint"],
        names=index_names,
        axis=0,
    ).stack()

    return ds


# %%
def calc_snac_sector_accounts(
    SLdom,
    QAimp_Ldom,
    Ydom_diag,
    Qimp_mrio,
    Yimp_diag,
    Ydom_exp_diag,
    Yimp_exp_diag,
    total_imports,
    pba,
):
    """Calculate and merge sector accounts

    Parameters
    ----------
    SLdom: pd.DataFrame

    QAimp_Ldom: pd.DataFrame

    Qimp_mrio: pd.DataFrame

    Ydom_diag: pd.DataFrame

    Yimp_diag: pd.DataFrame

    Ydom_exp_diag: pd.DataFrame

    Yimp_exp_diag: pd.DataFrame

    total_imports: pd.DataFrame

    pba: pd.DataFrame

    Returns
    --------
    pd.Series

        index: pd.MultiIndex

            level 0: component
                domestic_footprint, import_footprint, total_footprint
            level 1: emission
                sector names
            level 2: sector
                sector names
    """

    total_footprint = SLdom @ Ydom_diag + QAimp_Ldom @ Ydom_diag + Qimp_mrio @ Yimp_diag

    domestic_footprint = SLdom * Ydom_diag.sum(axis=1)
    import_footprint = QAimp_Ldom @ Ydom_diag + Qimp_mrio @ Yimp_diag
    export_footprint = (
        SLdom @ Ydom_exp_diag + QAimp_Ldom @ Ydom_exp_diag + Qimp_mrio @ Yimp_exp_diag
    )

    footprint_gross_imports = Qimp_mrio * total_imports
    footprint_gross_imports.columns.name = "sector"

    index_names = ["component", "emission"]

    ds = pd.concat(
        [
            total_footprint,
            domestic_footprint,
            import_footprint,
            export_footprint,
            footprint_gross_imports,
            pba,
        ],
        keys=[
            "total_footprint",
            "domestic_footprint",
            "import_footprint",
            "footprint_gross_exports",
            "footprint_gross_imports",
            "production_account",
        ],
        names=index_names,
        axis=0,
    ).stack()

    return ds


# %%
def exio_nor_converter(to_conv, sector_matching, new_col_names="name"):
    """Converts classification from EXIOBASE to Norwegian sectors

    This checks if the columns of to_conv matches on of
    the index of sector_matching.

    Parameters
    ----------
    to_conv: pd.DataFrame
        Accounts to convert to Norwegian sectors.
        EXIOBASE classification in the columns.

    sector_matching: pd.DataFrame
        Matching between EXIOBASE and Norwegian sectors.
        Read from sector matching file.

    new_col_names: str
        Can be 'name', 'code' or 'nr'.
        Taken from sector_matching columns and used as the new
        sector names.

    Returns
    -------

    pd.DataFrame
        Accounts converted to Norwegian sectors
    """
    # As we do an concordance based conversion, we check for the correct order
    # for at least on of the index levels of sector_matching
    for ind_number in range(0, 100):
        try:
            sec_match_order = sector_matching.index.get_level_values(ind_number)
            if sec_match_order.equals(to_conv.columns):
                break
        except IndexError:
            raise ValueError("Could not find matching index value do not match")

    ret = to_conv @ sector_matching.values
    ret.columns = sector_matching.columns.get_level_values(new_col_names).str.strip()
    return ret


# %%
def chem_sector_agg(df):
    """Aggregates the three chemical sectors to one sector.

    Due to confidentiality of data for the three sectors below in both Norwegian emissions and economic data,
    the data is aggregated to one sector, to be consistent with the Norwegian IO data
    NOTE: the original sector name for "Coke ..." has a whitespace at the end.
    We removed that whitespace during reading the files.

    Parameters
    ----------
    df: pd.DataFrame
        Accounts in Norwegian classification in the columns.
        Rows can be anything.

    Returns:
    --------
    pd.DataFrame
        Accounts with the three chemical sectors aggregated to one sector.
        Rest of the accounts/rows are unchanged.

    """
    chem_sectors_to_sum = [
        "Coke and refined petroleum products",
        "Chemicals and chemical products",
        "Basic pharmaceutical products and pharmaceutical preparations",
    ]

    chem_sector_agg = "Basic pharmaceutical products and pharmaceutical preparations"

    chem_sum = df.loc[:, chem_sectors_to_sum].sum(axis=1)
    df.loc[:, chem_sectors_to_sum] = 0
    df.loc[:, chem_sector_agg] = chem_sum
    return df


# %% [markdown]
# ## Main calculation loop

# %% [markdown]
# Establish some dictionaries for storing the results

# %%
col_dom_io = {}
col_Q_imp_exio = {}
col_snac = {}
col_findem_breakdown = {}
col_sector_accounts = {}
col_grand_totals = {}
col_sources = {}
col_sources_tot = {}

# %% [markdown]
# Run the calculation for each year

# %%
for year in years:
    print(f"PROCESSING YEAR: {year}")

    dom_io = get_nor_IO(
        file_dom=nor_io_dom_files[year],
        file_imp=nor_io_imp_files[year],
        emis_nor_df=emis_nor_ind[year],
    )

    imp_exio_extract = pd.read_excel(
        exio_extract_file, sheet_name=f"imp_{year}", index_col=0, header=[0, 1]
    ).T.reset_index("unit")
    imp_exio_extract_unit = imp_exio_extract.loc[:, "unit"].squeeze().copy()
    imp_exio_extract_unit.index.name = "account"
    imp_exio_extract_unit.name = "unit"
    imp_exio_extract = imp_exio_extract.drop("unit", axis=1)

    imp_nor = exio_nor_converter(
        imp_exio_extract, sector_matching, new_col_names="name"
    )
    imp_nor.loc["imports"] = (
        imp_nor.loc["imports"] * exchange_euro_nok.loc[year, "euro_to_nok"]
    )
    imp_nor_unit = imp_exio_extract_unit.copy()
    imp_nor_unit.loc["imports"] = "M.NOK"

    # Fix emission mismatch
    # Due to confidentiality of data for the three sectors below in both Norwegian emissions and economic data,
    # the data is aggregated to one sector, to be consistent with the Norwegian IO data
    imp_nor = chem_sector_agg(imp_nor)

    # We have Q_ something as abbreviation for the (full-supply-chain) multipliers
    # throughout the script. Note, that these are called .M in the pymrio package.
    Q_imp_exio = (
        imp_nor.div(imp_nor.loc["imports", :], axis=1).fillna(0).drop("imports", axis=0)
    )
    Q_imp_exio_unit = imp_nor_unit.drop("imports", axis=0)
    Q_imp_exio_unit = Q_imp_exio_unit + "/M.NOK"

    snac = snac_coupling(Aimp=dom_io["Aimp"], Ldom=dom_io["Ldom"], Qimp_mrio=Q_imp_exio)

    findem_breakdown_gather = dict()
    for emission in emis_nor_unit.index:
        SLdom_diag = pd.DataFrame(
            np.diag(dom_io["SLdom"].loc[emission, :]),
            index=dom_io["SLdom"].columns,
            columns=dom_io["SLdom"].columns,
        )

        QAimp_Ldom_diag = pd.DataFrame(
            np.diag(snac["QAimpLdom"].loc[emission, :]),
            index=snac["QAimpLdom"].columns,
            columns=snac["QAimpLdom"].columns,
        )
        Q_imp_exio_diag = pd.DataFrame(
            np.diag(Q_imp_exio.loc[emission, :]),
            index=Q_imp_exio.columns,
            columns=Q_imp_exio.columns,
        )
        findem_breakdown_gather[emission] = snac_findem_breakdown(
            SLdom_diag=SLdom_diag,
            QAimp_Ldom_diag=QAimp_Ldom_diag,
            Q_imp_exio_diag=Q_imp_exio_diag,
            Ydom_br=dom_io["Ydom_br"],
            Yimp_br=dom_io["Yimp_br"],
        )

    _df = pd.concat(findem_breakdown_gather, axis=1)
    _df.columns.names = ["emission"]

    fp_findem_breakdown = _df.stack()
    del _df

    # footprint of imports (goods traded at the border)
    total_imports = dom_io["Zimp"].sum(axis=1) + dom_io["Yimp"] + dom_io["Yimp_exp"]

    sector_accounts = calc_snac_sector_accounts(
        SLdom=dom_io["SLdom"],
        QAimp_Ldom=snac["QAimpLdom"],
        Ydom_diag=pd.DataFrame(
            data=np.diag(dom_io["Ydom"]),
            index=dom_io["Ydom"].index,
            columns=dom_io["Ydom"].index,
        ),
        Yimp_diag=pd.DataFrame(
            data=np.diag(dom_io["Yimp"]),
            index=dom_io["Yimp"].index,
            columns=dom_io["Yimp"].index,
        ),
        Ydom_exp_diag=pd.DataFrame(
            data=np.diag(dom_io["Ydom_exp"]),
            index=dom_io["Ydom_exp"].index,
            columns=dom_io["Ydom_exp"].index,
        ),
        Yimp_exp_diag=pd.DataFrame(
            data=np.diag(dom_io["Yimp_exp"]),
            index=dom_io["Yimp_exp"].index,
            columns=dom_io["Yimp_exp"].index,
        ),
        Qimp_mrio=Q_imp_exio,
        total_imports=total_imports,
        pba=emis_nor_ind[year],
    )

    # fp-src calc: get source of footprint from exiobase
    _source = pd.read_excel(
        exio_extract_file,
        sheet_name=f"imp_source_{year}",
        index_col=[0, 1],
        header=[0],
    )
    _source.columns.name = "region"
    _source = (
        _source.stack("region")
        .unstack("sector")
        .reindex(columns=sector_matching.index.get_level_values("code"))
    )

    tot_ghg = _source.loc[("GHG")].sum().sum()
    tot_ghg
    nor_ghg = _source.loc[("GHG", "NO")].sum()
    nor_ghg
    tot_ghg - nor_ghg

    # fp-src calc: convert to nor class and do the chem sector adjustment
    source_nor_class = exio_nor_converter(
        _source, sector_matching, new_col_names="name"
    )
    source_nor_class.columns.name = "sector"
    source_nor_class = chem_sector_agg(source_nor_class)

    # fp-src: get src scaled import mulitpliers for each region
    Q_src_imp = (
        source_nor_class.div(imp_nor.loc["imports", :], axis=1)
        .fillna(0)
        .replace(np.inf, 0)
    )

    # fp-src: multiply by imports and dom demand
    Q_Aimp_src = Q_src_imp @ dom_io["Aimp"]
    Q_Aimp_Ldom_src = Q_Aimp_src @ dom_io["Ldom"]

    # fp-src: multiply by final demand and add norwegian domestic demand
    src_all = Q_Aimp_Ldom_src * dom_io["Ydom"] + Q_src_imp * dom_io["Yimp"]
    src_all.loc[(slice(None), "NO"), :] = (
        src_all.loc[(slice(None), "NO"), :] + dom_io["SLdom"] * dom_io["Ydom"]
    )
    src_all = src_all.stack()
    src_all.name = "value"

    src_total = src_all.groupby(level=["region", "emission"]).sum()

    # calculate some convenient totals
    total_accounts = pd.DataFrame(
        sector_accounts.groupby(["component", "emission"]).sum(),
        columns=["value"],
    )
    total_hhld = pd.DataFrame(emis_nor_hhld[year].loc[:, "Households, totals"])
    total_hhld.columns = ["value"]
    tot_fp_with_hhld = (
        (total_accounts.loc["total_footprint"] + total_hhld)
        .assign(component="total_footprint_with_households")
        .set_index("component", append=True)
        .reorder_levels(["component", "emission"])
    )
    tot_pba_with_hhld = (
        (total_accounts.loc["production_account"] + total_hhld)
        .assign(component="production_account_with_households")
        .set_index("component", append=True)
        .reorder_levels(["component", "emission"])
    )
    total_hhld = (
        total_hhld.assign(component="household_totals")
        .set_index("component", append=True)
        .reorder_levels(["component", "emission"])
    )
    total_accounts = pd.concat(
        [total_accounts, tot_fp_with_hhld, tot_pba_with_hhld, total_hhld],
        axis=0,
    ).squeeze()

    col_dom_io[year] = dom_io
    col_Q_imp_exio[year] = Q_imp_exio
    col_snac[year] = snac
    col_findem_breakdown[year] = fp_findem_breakdown
    col_sector_accounts[year] = sector_accounts
    col_grand_totals[year] = total_accounts
    col_sources[year] = src_all
    col_sources_tot[year] = src_total

# %% [markdown]
# ## Preparing and storing the results
# For storing the results we stack them together in a long list format.
# This can easily be converted to any data format, including SQL.
# In excel, the list can easily be pivoted to a wide format or filtered for specific years.
# In python, this can be extended to a wide format with .unstack(index_name).
# We add the unit to the list already here, to make it easier to work with in excel.
# For subsequent calculations, the unit might be cumbersome
# and can be dropped with .drop('unit', axis=1)

# %% [markdown]
# Helper function for stacking, naming and unit assignment


# %%
def stack_and_unit(dd, emis_units, sector_order):
    """Stacks dataframes for final save and assign units

    The results are ordered
        - emission: order given in emis_units
        - sectors: order given in sector_order
        - years: ascending

    Parameters
    ----------
    dd : dict
        With years as keys and data-frames as values
    emis_units : pd.Series
        Emission units with index of emission names.
        The order of the index is used for the order of
        the result.
    sector_order : list
        Orders list of sectors for reordering the results

    Returns
    -------
    pd.DataFrame
        index: Multiindex with year and the remaining ones from the dd values
        columns: value, unit

    """

    merged = pd.concat(dd, axis=0, names=["year"])

    while type(merged) is pd.DataFrame:
        merged = merged.stack()

    merged.name = "value"
    _ix = merged.index
    df = merged.reset_index().merge(emis_nor_unit, on="emission").set_index(_ix.names)
    df = df.sort_index(level="year", sort_remaining=False).reindex(
        emis_units.index.tolist(), level="emission"
    )

    if "sector" in df.index.names:
        df = df.reindex(sort_sector, level="sector")

    return df


# %% [markdown]
# Stacking the results into a long list format

# %%
sort_sector = dom_io["x"].index.tolist()

all_sector_accounts = stack_and_unit(
    col_sector_accounts, emis_units=emis_nor_unit, sector_order=sort_sector
)
all_findem_breakdown = stack_and_unit(
    col_findem_breakdown, emis_units=emis_nor_unit, sector_order=sort_sector
)
all_hhld = stack_and_unit(
    emis_nor_hhld, emis_units=emis_nor_unit, sector_order=sort_sector
)
all_sources = stack_and_unit(
    col_sources, emis_units=emis_nor_unit, sector_order=sort_sector
)
all_sources_tot = stack_and_unit(
    col_sources_tot, emis_units=emis_nor_unit, sector_order=sort_sector
)
all_totals = stack_and_unit(
    col_grand_totals, emis_units=emis_nor_unit, sector_order=sort_sector
)

# %% [markdown]
# Finally we store the results

# %%
all_sector_accounts.to_csv(output_path / "sector_accounts.tsv", sep="\t")
all_findem_breakdown.to_csv(
    output_path / "footprints_final_demand_breakdown.tsv", sep="\t"
)
all_hhld.to_csv(output_path / "household_emissions.tsv", sep="\t")
all_totals.to_csv(output_path / "total_accounts.tsv", sep="\t")
all_sources.to_csv(output_path / "footprint_sources.tsv", sep="\t")
all_sources_tot.to_csv(output_path / "footprint_sources_totals.tsv", sep="\t")

# %%
print(f"Done - results stored at {output_path}")
